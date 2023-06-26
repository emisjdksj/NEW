import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import load_workbook

colors = ['#7F8E95', '#90C9E6', '#269EBC', '#023048', '#C43B3F', '#7F8E95']    # 6 bins max default

path = '0626.xlsx'
fig_path = 'pie.png'
df = pd.read_excel(path)
attributes = df.columns.tolist()[2:]
million = 1e6
bins_labels = {
    'Beta': {'bins': [0, 0.2, 1, float('inf')], 'labels': ['0-0.2', '0.2-1', '> 1']},
    'Liquidity': {'bins': [0, 5*million, 10*million, float('inf')],
                  'labels': ['0-5(million)', '5-10(million)', '> 10(million)']},
    'Vol': {'bins': [0, 10, 20, float('inf')], 'labels': ['0-10', '10-20', '> 20']}
}

for ind, attr in enumerate(attributes):
    plt.figure(figsize=(6, 6))
    data = df[attr]
    if attr in ['Beta', 'Liquidity', 'Vol']:
        data = pd.cut(df[attr], bins=bins_labels[attr]['bins'], right=False, labels=bins_labels[attr]['labels'])
    data = pd.value_counts(data, sort=True, normalize=True)
    labels = data.index
    # plt.subplot(231+ind)
    plt.pie(data, labels=labels, colors=colors[:len(labels)], autopct='%1.1f%%', shadow=False,
            textprops={'fontsize': 12, 'color': 'w'})
    plt.title(attr)
    plt.legend(frameon=False)
    plt.savefig(attr+'.png', bbox_inches='tight')

wb = load_workbook(path)
for attr in attributes:
    pie = wb.create_sheet(attr)
    pie.add_image(openpyxl.drawing.image.Image(attr+'.png'), 'A1')


# # subplots
# plt.figure(figsize=(15, 8))
# for ind, attr in enumerate(attributes):
#     data = df[attr]
#     if attr in ['Beta', 'Liquidity', 'Vol']:
#         data = pd.cut(df[attr], bins=bins_labels[attr]['bins'], right=False, labels=bins_labels[attr]['labels'])
#     data = pd.value_counts(data, sort=True, normalize=True)
#     labels = data.index
#     plt.subplot(231+ind)
#     plt.pie(data, labels=labels, colors=colors[:len(labels)], autopct='%1.1f%%')
#     plt.title(attr)
# plt.savefig(fig_path, bbox_inches='tight')
#
# wb = load_workbook(path)
# pie = wb.create_sheet('Pie')
# pie.add_image(openpyxl.drawing.image.Image(fig_path), 'A1')

wb.save(path)
