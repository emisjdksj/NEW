
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os
from openpyxl.utils import get_column_letter, column_index_from_string

path = 'OP balance.xls'
tgt_path = r'./工作簿1.xlsx'
convert2billion = True
objs = ['A', 'B']
width = 0.4
colors = ['#7F8E95', '#90C9E6']

# Read the Excel file into a DataFrame
df = pd.read_excel(path)

# Convert the 'Date' column to datetime type
df['date'] = pd.to_datetime(df['date'], format='%m-%d')

# Calculate the total notional amount for each day
total_op_market_day = df[df['market'].isin(objs)].groupby(['date', 'market'])['OP balance'].sum().reset_index()
total_net_market_day = df[df['market'].isin(objs)].groupby(['date', 'market'])['net'].sum().reset_index()

total_op_market_day = total_op_market_day.pivot(index='date', columns='market', values='OP balance').fillna(0)
total_net_market_day = total_net_market_day.pivot(index='date', columns='market', values='net').fillna(0)

fig, ax1 = plt.subplots(figsize=(10, 6))
labels = total_op_market_day.index.tolist()  # x label (date)
market_names = total_op_market_day.columns.tolist()  # market (==objs)

x = np.arange(len(labels))
markets_op_data = []
billion = 1
max_op = -float('inf')
for i, col in enumerate(total_op_market_day.columns):
    markets_op_data.append((total_op_market_day[col]/billion).tolist())
    max_op = max(max_op, max(markets_op_data[-1]))
if convert2billion:
    plt.text(0. if max_op/1e9 > 1 else .055, 1.01, s='billion', transform=ax1.transAxes)
    billion = 1e9
markets_op_data = []
for i, col in enumerate(total_op_market_day.columns):
    markets_op_data.append((total_op_market_day[col]/billion).tolist())
for ind, c in enumerate(markets_op_data):
    ax1.bar(x-width/2+ind*width, c, width, label=market_names[ind], color=colors[ind], alpha=.8)

ax1.set_ylabel('Total OP Balance Amount')
ax1.set_xticks(x)
ax1.set_xticklabels(labels)
xtick_labels = [x.strftime('%m-%d') for x in labels]
ax1.set_xticklabels(xtick_labels, rotation=30, ha='center')

ax2 = ax1.twinx()
markets_net_data = []
billion = 1e9 if convert2billion else 1
max_net = -float('inf')
for col in total_net_market_day.columns:
    markets_net_data.append((total_net_market_day[col]/billion).tolist())
for ind, c in enumerate(markets_net_data):
    ax2.plot(x, c, color=colors[ind], marker='o', linestyle='dashed')
    max_net = max(max_net, max(c))
ax2.set_ylabel('Total Net Amount')
ax1.set_xlabel('Position Date')
ax1.legend(loc='upper left', frameon=False)
fig.tight_layout()
plt.savefig('opbalance1.png')
# plt.show()

# overview
total_net_day, total_op_day = df.groupby('date')['net'].sum().reset_index(), df.groupby('date')['OP balance'].sum().reset_index()
date_range = total_net_day['date']
df_daily = pd.DataFrame({'date': date_range})

df_merged = pd.merge(df_daily, total_net_day, on='date', how='left')
df_merged = pd.merge(df_merged, total_op_day, on='date', how='left')
df_merged.fillna(0, inplace=True)
fig, ax1 = plt.subplots(figsize=(10, 6))

billion = 1e9 if convert2billion else 1
if convert2billion:
    plt.text(0, 1.01, s='billion', transform=ax1.transAxes)
x = np.arange(len(df_merged['date'].tolist()))
labels = df_merged['date'].tolist()
ax1.bar(x, (df_merged['OP balance']/billion).tolist(), color='lightblue', width=0.6)
ax1.set_xlabel('Position Date')
ax1.set_ylabel('Total OP Balance Amount')
ax1.set_xticks(x)
ax1.tick_params(axis='y')

# Creating a secondary axis for the weighted average spread
ax2 = ax1.twinx()
ax2.plot(x, df_merged['net']/billion, color='darkblue', marker='o')
ax2.set_ylabel('Total Net Amount')
ax2.tick_params(axis='y')

xtick_labels = [x.strftime('%m-%d') for x in labels]
ax1.set_xticklabels(xtick_labels, rotation=30, ha='center')
# Adjusting the layout with increased spacing
plt.subplots_adjust(left=0.1, right=0.9, top=0.9, bottom=0.1)

# Save the chart as an image
plt.savefig('./opbalance2.png')
# plt.show()

wb = load_workbook(tgt_path)
temp = wb.create_sheet('OP balance (AB)')
temp.add_image(openpyxl.drawing.image.Image('./opbalance1.png'), 'A1')
temp = wb.create_sheet('OP balance (Total)')
temp.add_image(openpyxl.drawing.image.Image('./opbalance2.png'), 'A1')

# Save the modified Excel file
wb.save(tgt_path)
