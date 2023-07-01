import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import load_workbook
from matplotlib import rcParams

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string

config = {
    "font.family": 'serif',
    # "font.size": 14,
    "mathtext.fontset": 'stix',
}
rcParams.update(config)

path = 'OP balance.xls'
tgt_path = r'./工作簿1.xlsx'
convert2billion = True
objs = ['A', 'B']
width = 0.4
colors = ['#7F8E95', '#90C9E6']

wb = load_workbook(tgt_path)
# Read the Excel file into a DataFrame
df = pd.read_excel(path)

# Convert the 'Date' column to datetime type
df['date'] = pd.to_datetime(df['date'], format='%m-%d')

# Calculate the total notional amount for each day
# sum 如果market每天仅有一行数据，该计算不影响结果正确性
total_op_market_day = df[df['market'].isin(objs)].groupby(['date', 'market'])['OP balance'].sum().reset_index()
total_net_market_day = df[df['market'].isin(objs)].groupby(['date', 'market'])['net'].sum().reset_index()
total_utility_market_day = df[df['market'].isin(objs)].groupby(['date', 'market'])['utility'].sum().reset_index()

total_op_market_day = total_op_market_day.pivot(index='date', columns='market', values='OP balance').fillna(0)
total_net_market_day = total_net_market_day.pivot(index='date', columns='market', values='net').fillna(0)
total_utility_market_day = total_utility_market_day.pivot(index='date', columns='market', values='utility').fillna(0)

labels = total_op_market_day.index.tolist()  # x label (date)
market_names = total_op_market_day.columns.tolist()  # market (==objs)

# A
fig, ax1 = plt.subplots(figsize=(10, 6))
x = np.arange(len(labels))
markets_op_data = []
billion = 1
max_op = -float('inf')
for i, col in enumerate(total_op_market_day.columns):
    markets_op_data.append((total_op_market_day[col]/billion).tolist())
    max_op = max(max_op, max(markets_op_data[-1]))
if convert2billion:
    plt.text(0. if max_op/1e9 > 1 else .055, 1.01, s='billion', transform=ax1.transAxes)
    billion = 1e9
markets_op_data = []
for i, col in enumerate(total_op_market_day.columns):
    markets_op_data.append((total_op_market_day[col]/billion).tolist())
# for ind, c in enumerate(markets_op_data):
#     ax1.bar(x-width/2+ind*width, c, width, label=market_names[ind], color=colors[ind], alpha=.8)

ax1.bar(x, markets_op_data[0], width, color='lightblue', alpha=.8, label='Total OP balance amount')  # 0630 图例标签
ax1.set_ylabel('Total OP Balance Amount')
ax1.set_xticks(x)
ax1.set_xticklabels(labels)
xtick_labels = [x.strftime('%m-%d') for x in labels]
ax1.set_xticklabels(xtick_labels, rotation=30, ha='center')

ax2 = ax1.twinx()
markets_net_data = []
billion = 1e9 if convert2billion else 1
for col in total_net_market_day.columns:
    markets_net_data.append((total_net_market_day[col]).tolist())
# for ind, c in enumerate(markets_net_data):
#     ax2.plot(x, c, color=colors[ind], marker='o', linestyle='dashed')
ax2.plot(x, markets_net_data[0], marker='o', color='darkblue', label='Total net amount')  # 0630 图例标签
ax2.set_ylabel('Total Net Amount')
ax1.set_xlabel('Position Date')
# ax1.legend(loc='upper left', frameon=False, bbox_to_anchor=(.42, -.15), ncol=2)
ax1.set_title('A')
ax1.legend(frameon=False, bbox_to_anchor=(.5, -.14))  # 0630 图例显示
ax2.legend(frameon=False, bbox_to_anchor=(.72, -.14))  # 0630 图例显示
fig.tight_layout()
plt.savefig('opbalance_A.png')


"""
write table
"""
ws = wb.create_sheet('OP balance (A)')
# Define the cell style
title_font_bold = Font(bold=True, color='ffffff')
font_bold = Font(bold=True)
alignment_center = Alignment(horizontal='center', vertical='center')
border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                     bottom=Side(style='thin'))
fill = PatternFill(start_color="90C9E6", fill_type='solid')

# Write the client difference table to the worksheet
ws['A1'] = 'Position Date'
ws['A1'].font = title_font_bold
ws['A1'].alignment = alignment_center
ws['A1'].border = border_thin
ws['A1'].fill = fill

ws['B1'] = 'Total OP Balance Amount'
ws['B1'].font = title_font_bold
ws['B1'].alignment = alignment_center
ws['B1'].border = border_thin
ws['B1'].fill = fill

ws['C1'] = 'Total Net Amount'
ws['C1'].font = title_font_bold
ws['C1'].alignment = alignment_center
ws['C1'].border = border_thin
ws['C1'].fill = fill
for i in range(len(total_op_market_day.index.tolist())):
    row = i + 2
    ws[f'A{row}'] = str(total_op_market_day.index.tolist()[i]).split()[0]
    ws[f'A{row}'].border = border_thin

    ws[f'B{row}'] = total_op_market_day.iloc[i]['A']
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].number_format = '#,##0'

    ws[f'C{row}'] = total_net_market_day.iloc[i]['A']
    ws[f'C{row}'].border = border_thin
    ws[f'C{row}'].number_format = '#,##0'

for col in ws.columns:
    max_length = 0
    column = col[0]
    if column.data_type == 's':
        column = get_column_letter(column.column)
    else:
        column = get_column_letter(column_index_from_string(column.coordinate[:1]))
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = max_length + 2
    ws.column_dimensions[column].width = adjusted_width

ws.add_image(openpyxl.drawing.image.Image('./opbalance_A.png'), 'E1')

# B
fig, ax1 = plt.subplots(figsize=(10, 6))
if convert2billion:
    plt.text(0. if max_op/1e9 > 1 else .055, 1.01, s='billion', transform=ax1.transAxes)
    billion = 1e9
ax1.bar(x, markets_op_data[1], width, color='lightblue', alpha=.8, label='Total OP balance amount')  # 0630 图例标签
ax1.set_ylabel('Total OP Balance Amount')
ax1.set_xticks(x)
ax1.set_xticklabels(labels)
xtick_labels = [x.strftime('%m-%d') for x in labels]
ax1.set_xticklabels(xtick_labels, rotation=30, ha='center')

ax2 = ax1.twinx()
markets_net_data = []
billion = 1e9 if convert2billion else 1
for col in total_net_market_day.columns:
    markets_net_data.append((total_net_market_day[col]).tolist())
# for ind, c in enumerate(markets_net_data):
#     ax2.plot(x, c, color=colors[ind], marker='o', linestyle='dashed')
ax2.plot(x, markets_net_data[1], marker='o', color='darkblue', label='Total net amount')  # 0630 图例标签
ax2.set_ylabel('Total Net Amount')
ax1.set_xlabel('Position Date')
# ax1.legend(loc='upper left', frameon=False, bbox_to_anchor=(.42, -.15), ncol=2)
ax1.set_title('B')
ax1.legend(frameon=False, bbox_to_anchor=(.5, -.14))  # 0630 图例显示
ax2.legend(frameon=False, bbox_to_anchor=(.72, -.14))  # 0630 图例显示
fig.tight_layout()
plt.savefig('opbalance_B.png')
# plt.show()


"""
write table
"""
ws = wb.create_sheet('OP balance (B)')
# Define the cell style
title_font_bold = Font(bold=True, color='ffffff')
font_bold = Font(bold=True)
alignment_center = Alignment(horizontal='center', vertical='center')
border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                     bottom=Side(style='thin'))
fill = PatternFill(start_color="90C9E6", fill_type='solid')

# Write the client difference table to the worksheet
ws['A1'] = 'Position Date'
ws['A1'].font = title_font_bold
ws['A1'].alignment = alignment_center
ws['A1'].border = border_thin
ws['A1'].fill = fill

ws['B1'] = 'Total OP Balance Amount'
ws['B1'].font = title_font_bold
ws['B1'].alignment = alignment_center
ws['B1'].border = border_thin
ws['B1'].fill = fill

ws['C1'] = 'Total Net Amount'
ws['C1'].font = title_font_bold
ws['C1'].alignment = alignment_center
ws['C1'].border = border_thin
ws['C1'].fill = fill
for i in range(len(total_op_market_day.index.tolist())):
    row = i + 2
    ws[f'A{row}'] = str(total_op_market_day.index.tolist()[i]).split()[0]
    ws[f'A{row}'].border = border_thin

    ws[f'B{row}'] = total_op_market_day.iloc[i]['B']
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].number_format = '#,##0'

    ws[f'C{row}'] = total_net_market_day.iloc[i]['B']
    ws[f'C{row}'].border = border_thin
    ws[f'C{row}'].number_format = '#,##0'

for col in ws.columns:
    max_length = 0
    column = col[0]
    if column.data_type == 's':
        column = get_column_letter(column.column)
    else:
        column = get_column_letter(column_index_from_string(column.coordinate[:1]))
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = max_length + 2
    ws.column_dimensions[column].width = adjusted_width
ws.add_image(openpyxl.drawing.image.Image('./opbalance_B.png'), 'E1')

# overview
total_net_day, total_op_day = df.groupby('date')['utility'].sum().reset_index(), df.groupby('date')['OP balance'].sum().reset_index()
total_net_day['utility'] /= total_op_day['OP balance']
total_net_day['net'] = total_net_day['utility']
"""
total_net = (A.utility+B.utility)/(A.op+B.op)
"""

date_range = total_net_day['date']
df_daily = pd.DataFrame({'date': date_range})

df_merged = pd.merge(df_daily, total_net_day, on='date', how='left')
df_merged = pd.merge(df_merged, total_op_day, on='date', how='left')
df_merged.fillna(0, inplace=True)
fig, ax1 = plt.subplots(figsize=(10, 6))

billion = 1e9 if convert2billion else 1
if convert2billion:
    plt.text(0, 1.01, s='billion', transform=ax1.transAxes)
x = np.arange(len(df_merged['date'].tolist()))
labels = df_merged['date'].tolist()
ax1.bar(x, (df_merged['OP balance']/billion).tolist(), color='lightblue', width=0.6, label='Total OP balance amount')  # 0630 图例标签
ax1.set_xlabel('Position Date')
ax1.set_ylabel('Total OP Balance Amount')
ax1.set_xticks(x)
ax1.tick_params(axis='y')

# Creating a secondary axis for the weighted average spread
ax2 = ax1.twinx()
ax2.plot(x, df_merged['net'], color='darkblue', marker='o', label='Total net amount')  # 0630 图例标签
ax2.set_ylabel('Total Net Amount')
ax2.tick_params(axis='y')
ax1.set_ylim(-100, 20)
xtick_labels = [x.strftime('%m-%d') for x in labels]
ax1.set_xticklabels(xtick_labels, rotation=30, ha='center')
# Adjusting the layout with increased spacing
# plt.subplots_adjust(left=0.1, right=0.9, top=0.9, bottom=0.1)
plt.title('In Total')
ax1.legend(frameon=False, bbox_to_anchor=(.5, -.14))  # 0630 图例显示
ax2.legend(frameon=False, bbox_to_anchor=(.72, -.14))  # 0630 图例显示
fig.tight_layout()
# Save the chart as an image
plt.savefig('./opbalance.png')
# plt.show()


"""
write table
"""
ws = wb.create_sheet('OP balance (Total)')
# Define the cell style
title_font_bold = Font(bold=True, color='ffffff')
font_bold = Font(bold=True)
alignment_center = Alignment(horizontal='center', vertical='center')
border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                     bottom=Side(style='thin'))
fill = PatternFill(start_color="90C9E6", fill_type='solid')

# Write the client difference table to the worksheet
ws['A1'] = 'Position Date'
ws['A1'].font = title_font_bold
ws['A1'].alignment = alignment_center
ws['A1'].border = border_thin
ws['A1'].fill = fill

ws['B1'] = 'Total OP Balance Amount'
ws['B1'].font = title_font_bold
ws['B1'].alignment = alignment_center
ws['B1'].border = border_thin
ws['B1'].fill = fill

ws['C1'] = 'Total Net Amount'
ws['C1'].font = title_font_bold
ws['C1'].alignment = alignment_center
ws['C1'].border = border_thin
ws['C1'].fill = fill
for i in range(len(total_op_market_day.index.tolist())):
    row = i + 2
    ws[f'A{row}'] = str(df_merged.iloc[i]['date']).split()[0]
    ws[f'A{row}'].border = border_thin

    ws[f'B{row}'] = df_merged.iloc[i]['OP balance']
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].number_format = '#,##0'       # or #,##0.00

    ws[f'C{row}'] = df_merged.iloc[i]['net']
    ws[f'C{row}'].border = border_thin
    ws[f'C{row}'].number_format = '#,##0'

for col in ws.columns:
    max_length = 0
    column = col[0]
    if column.data_type == 's':
        column = get_column_letter(column.column)
    else:
        column = get_column_letter(column_index_from_string(column.coordinate[:1]))
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = max_length + 2
    ws.column_dimensions[column].width = adjusted_width
ws.add_image(openpyxl.drawing.image.Image('./opbalance.png'), 'E1')

# Save the modified Excel file
wb.save(tgt_path)
