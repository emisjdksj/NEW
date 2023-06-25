
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os
from openpyxl.utils import get_column_letter, column_index_from_string

pd.set_option('display.max_columns', 1000)
pd.set_option('display.width', 1000)
pd.set_option('display.max_colwidth', 1000)
pd.set_option('display.unicode.ambiguous_as_wide', True)
pd.set_option('display.unicode.east_asian_width', True)

# Setting up the path to the Excel file
path = r'C:\Users\yiyin\Desktop\工作簿1.xlsx'

# Read the Excel file into a DataFrame
df = pd.read_excel(path)

# Convert the 'Date' column to datetime type
df['position date'] = pd.to_datetime(df['position date'], format='%m-%d')

# Calculate the total notional amount for each day
total_notional_day = df.groupby('position date')['axe notional mtm'].sum().reset_index()

# Calculate the weighted spread for each day
weighted_spread = df.groupby('position date').apply(lambda x: np.average(x['spread'], weights=x['axe notional mtm'])).reset_index()
weighted_spread.columns = ['position date', 'weighted_spread']
# print(weighted_spread)
# print(weighted_spread['position date'])
# exit(0)
# Create a new DataFrame with daily date range
# date_range = pd.date_range(start=total_notional_day['position date'].min(), end=total_notional_day['position date'].max(), freq='D')
date_range = weighted_spread['position date']
df_daily = pd.DataFrame({'position date': date_range})


# Merge the daily DataFrame with the total notional DataFrame
df_merged = pd.merge(df_daily, total_notional_day, on='position date', how='left')

# Merge the weighted_spread DataFrame with the df_merged DataFrame
df_merged = pd.merge(df_merged, weighted_spread, on='position date', how='left')
# fillna to replace nan values
df_merged.fillna(0, inplace=True)

# Plotting the chart with adjusted parameters
fig, ax1 = plt.subplots(figsize=(10, 6))
convert_million = True
million = 1
if convert_million:
    plt.text(0, 1.01, s='million', transform=ax1.transAxes)
    million = 1e6
# Plotting the total notional amount as a bar chart with adjusted width
x = np.arange(len(df_merged['position date'].tolist()))
labels = df_merged['position date'].tolist()
ax1.bar(x, (df_merged['axe notional mtm']/million).tolist(), color='lightblue', width=0.6)
ax1.set_xlabel('Position Date')
ax1.set_ylabel('Total Notional Amount')
ax1.set_xticks(x)
ax1.tick_params(axis='y')

# Creating a secondary axis for the weighted average spread
ax2 = ax1.twinx()
ax2.plot(x, df_merged['weighted_spread'], color='darkblue', marker='o')
ax2.set_ylabel('Weighted Average Spread')
ax2.tick_params(axis='y')

xtick_labels = [x.strftime('%m-%d') for x in labels]
ax1.set_xticklabels(xtick_labels, rotation=30, ha='center')
# Adjusting the layout with increased spacing
plt.subplots_adjust(left=0.1, right=0.9, top=0.9, bottom=0.1)

# Save the chart as an image
chart_path = 'chart.png'
plt.savefig(chart_path)

# Load the workbook
wb = load_workbook(path)

# Create a new worksheet named "Overview"
ws = wb.create_sheet('Overview')

# Add the chart image to the new worksheet
img = openpyxl.drawing.image.Image(chart_path.replace('\\', '/'))
img.anchor = 'A1'
ws.add_image(img)

# Calculate the total notional amount for each client on each day
total_notional = df.groupby(['position date', 'client'])['axe notional mtm'].sum().reset_index()

# Calculate the first and last notional for each client
client_first_notional = total_notional.groupby('client').first().reset_index()
client_last_notional = total_notional.groupby('client').last().reset_index()

# Calculate the notional difference for each client
client_diff = pd.merge(client_first_notional, client_last_notional, on='client', suffixes=('_first', '_last'))
client_diff['Notional Difference'] = client_diff['axe notional mtm_last'] - client_diff['axe notional mtm_first']

# Select only clients with Notional Difference > 0
client_diff = client_diff[client_diff['Notional Difference'] > 0]

# Select top 5 clients with largest absolute notional difference
client_diff['Abs Notional Difference'] = client_diff['Notional Difference'].abs()
client_diff = client_diff.nlargest(5, 'Abs Notional Difference')

# Calculate the overall notional difference
overall_diff = df_merged['axe notional mtm'].iloc[-1] - df_merged['axe notional mtm'].iloc[0]

# Calculate the difference proportion for each client
client_diff['Difference Proportion'] = client_diff['Notional Difference'] / overall_diff

# Sort and rank the client differences
client_diff = client_diff.sort_values('Notional Difference', ascending=False).reset_index(drop=True)
# Show the client difference
print(client_diff[['client', 'Notional Difference', 'Difference Proportion']])

# Create a new worksheet for the client difference table
ws_diff = wb.create_sheet('Client Difference')

# Define the cell style
font_bold = Font(bold=True)
alignment_center = Alignment(horizontal='center', vertical='center')
border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Write the client difference table to the worksheet
ws_diff['A1'] = 'Top 5 Clients with Largest Notional Difference'
ws_diff['A1'].font = font_bold
ws_diff.merge_cells('A1:C1')
ws_diff['A1'].alignment = alignment_center
ws_diff['A1'].border = border_thin

ws_diff['A2'] = 'Client'
ws_diff['A2'].font = font_bold
ws_diff['A2'].alignment = alignment_center
ws_diff['A2'].border = border_thin

ws_diff['B2'] = 'Notional Difference'
ws_diff['B2'].font = font_bold
ws_diff['B2'].alignment = alignment_center
ws_diff['B2'].border = border_thin

ws_diff['C2'] = 'Difference Proportion'
ws_diff['C2'].font = font_bold
ws_diff['C2'].alignment = alignment_center
ws_diff['C2'].border = border_thin

for i in range(len(client_diff)):
    row = i + 3
    ws_diff[f'A{row}'] = client_diff.iloc[i]['client']
    ws_diff[f'A{row}'].border = border_thin

    ws_diff[f'B{row}'] = client_diff.iloc[i]['Notional Difference']
    ws_diff[f'B{row}'].border = border_thin
    ws_diff[f'B{row}'].number_format = '#,##0'
    
    ws_diff[f'C{row}'] = str(client_diff.iloc[i]['Difference Proportion']*100)+'%'
    ws_diff[f'C{row}'].border = border_thin

for col in ws_diff.columns:
    max_length = 0
    column = col[0]
    if column.data_type == 's':
        column = get_column_letter(column.column)
    else:
        column = get_column_letter(column_index_from_string(column.coordinate[:1]))
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = max_length + 2
    ws_diff.column_dimensions[column].width = adjusted_width


# Create a new worksheet named "Top 5 Client (Change)"
ws_change = wb.create_sheet('Top 5 Client (Change)')

# Select the top 5 clients with largest notional difference
top5_clients = client_diff['client'].tolist()
# Filter the original DataFrame by the top 5 clients
df_top5 = df[df['client'].isin(top5_clients)]
top5_total_notional = df_top5.groupby(['position date', 'client'])['axe notional mtm'].sum().reset_index()
top5_weighted_spread = df_top5.groupby(['position date', 'client']).apply(lambda x: np.average(x['spread'], weights=x['axe notional mtm'])).reset_index()
top5_weighted_spread.columns = ['position date', 'client', 'weighted_spread']


# 将总名义金额和加权利差进行透视，创建一个表格
top5_notional_table = top5_total_notional.pivot(index='position date', columns='client', values='axe notional mtm').fillna(0)
top5_spread_table = top5_weighted_spread.pivot(index='position date', columns='client', values='weighted_spread').fillna(0)


# 绘制第一个组合图表并设置参数
fig1, ax1 = plt.subplots(figsize=(10, 6))

# 设置每个客户的颜色
# colors = ['#7F8E95', '#90C9E6', '#269EBC', '#136784', '#023048']      # v1
colors = ['#90C9E6', '#269EBC', '#023048', '#C43B3F', '#7F8E95']      # v2

# 绘制每个客户的总名义金额
width = 0.16
labels = top5_notional_table.index.tolist()  # x label
top5_clients_name = top5_notional_table.columns.tolist()  # clients
x = np.arange(len(labels))
clients_notional_data = []
million = 1
if convert_million:
    plt.text(0, 1.01, s='million', transform=ax1.transAxes)
    million = 1e6

for i, col in enumerate(top5_notional_table.columns):
    clients_notional_data.append((top5_notional_table[col]/million).tolist())
for ind, c in enumerate(clients_notional_data):
    ax1.bar(x-.32+ind*.16, c, width, label=top5_clients_name[ind], color=colors[ind])

# 设置左y轴的标签
ax1.set_ylabel('Total Notional Amount')
ax1.set_xticks(x)
ax1.set_xticklabels(labels)

xtick_labels = [x.strftime('%m-%d') for x in labels]
ax1.set_xticklabels(xtick_labels, rotation=30, ha='center')
# 创建第二个轴，用于加权平均利差
ax2 = ax1.twinx()
clients_spread_data = []
# 绘制每个客户的加权平均利差
max_spread = -float('inf')
for col in top5_spread_table.columns:
    clients_spread_data.append(top5_spread_table[col].tolist())
for ind, c in enumerate(clients_spread_data):
    ax2.plot(x, c, label=top5_clients_name[ind], color=colors[ind], marker='o', linestyle='dashed')
    max_spread = max(max_spread, max(c))
ax2.set_ylim(-100, (max_spread+100) // 100*100)

# 设置右y轴的标签
ax2.set_ylabel('Weighted Average Spread')

# 设置图表标题和标签
ax1.set_title('Top 5 Clients with Largest Notional Difference')
ax1.set_xlabel('Position Date')

# 设置图例
ax1.legend(loc='upper left', frameon=False)

# 调整图表布局
fig1.tight_layout()

# 保存第一个组合图表为图片
chart_path1 = 'chart1.png'
plt.savefig(chart_path1)

# 在Excel工作表中插入第一个组合图表
ws_change.add_image(openpyxl.drawing.image.Image(chart_path1), 'A1')

# 删除第一个组合图表的临时文件
# os.remove(chart_path1)


# Create a new worksheet named "Top 5 Client (Absolute)"
ws_absolute = wb.create_sheet('Top 5 Client (Absolute)')

# Select the top 5 clients with largest absolute total notional amount on the last day
# 选择最后一天名义金额最大的前5个客户
top5_absolute = df[df['position date'] == df['position date'].max()].groupby('client')['axe notional mtm'].sum().nlargest(5).reset_index()

# 将名义金额和加权利差进行透视，创建一个表格
df_top5 = df[df['client'].isin(top5_absolute['client'])]
absolute_notional_table = df_top5.pivot_table(index='position date', columns='client', values='axe notional mtm').fillna(0)
absolute_spread_table = df_top5.pivot_table(index='position date', columns='client', values='spread', aggfunc=lambda x: np.average(x, weights=df_top5.loc[x.index, 'axe notional mtm'])).fillna(0)

# 绘制第二个组合图表并设置参数
fig2, ax3 = plt.subplots(figsize=(10, 6))

# 设置每个客户的颜色
# colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd']

# 绘制每个客户的名义金额
width = 0.16
labels = absolute_notional_table.index.tolist()  # x label
top5_clients_name = absolute_notional_table.columns.tolist()  # clients

x = np.arange(len(labels))
clients_notional_data = []
million = 1
if convert_million:
    plt.text(0, 1.01, s='million', transform=ax3.transAxes)
    million = 1e6

for i, col in enumerate(absolute_notional_table.columns):
    clients_notional_data.append((absolute_notional_table[col]/million).tolist())
for ind, c in enumerate(clients_notional_data):
    ax3.bar(x-.32+ind*.16, c, width, label=top5_clients_name[ind], color=colors[ind])


# 设置左y轴的标签
ax3.set_ylabel('Notional Amount')

# 调整x轴刻度标签
xtick_labels = [x.strftime('%m-%d') for x in absolute_notional_table.index]
ax3.set_xticks(x)
ax3.set_xticklabels(xtick_labels, rotation=30, ha='center')

ax4 = ax3.twinx()

clients_spread_data = []
max_spread = -float('inf')
for col in absolute_spread_table.columns:
    clients_spread_data.append(absolute_spread_table[col].tolist())
for ind, c in enumerate(clients_spread_data):
    ax4.plot(x, c, label=top5_clients_name[ind], color=colors[ind], marker='o', linestyle='dashed')
    max_spread = max(max_spread, max(c))
ax4.set_ylim(-100, (max_spread+100) // 100*100)
ax4.set_ylabel('Weighted Average Spread')

# 设置图表标题和标签
ax3.set_title('Top 5 Clients with Largest Notional Amount on Latest Day')
ax3.set_xlabel('Position Date')

# 设置图例
ax3.legend(loc='upper left', frameon=False)

# 调整图表布局
fig2.tight_layout()

# 保存第二个组合图表为图片
chart_path2 = 'chart2.png'
plt.savefig(chart_path2)

# 在Excel工作表中插入第二个组合图表
ws_absolute.add_image(openpyxl.drawing.image.Image(chart_path2), 'A1')

# 删除第二个组合图表的临时文件
# os.remove(chart_path2)
    
# Save the modified Excel file
wb.save(path)
