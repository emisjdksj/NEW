import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import load_workbook
from matplotlib import rcParams

title_font_size, legend_font_size = 15, 14
xy_label_font_size = 13
xy_ticks_font_size = 13
pie_font_size = 12
config = {
    "font.family": 'serif',
    "font.size": title_font_size,
    "mathtext.fontset": 'stix',
}
rcParams.update(config)

colors = ['#7F8E95', '#90C9E6', '#269EBC', '#023048', '#C43B3F', '#7F8E95']  # 6 bins max default

path = '0626.xlsx'
fig_path = 'pie.png'
df = pd.read_excel(path)
attributes = df.columns.tolist()[2:]
million = 1e6
bins_labels = {
    'Beta': {'bins': [0, 0.2, 1, float('inf')], 'labels': ['0-0.2', '0.2-1', '> 1']},
    'Liquidity': {'bins': [0, 5 * million, 10 * million, float('inf')],
                  'labels': ['0-5(million)', '5-10(million)', '> 10(million)']},
    'Vol': {'bins': [0, 10, 20, float('inf')], 'labels': ['0-10', '10-20', '> 20']}
}

# for ind, attr in enumerate(attributes):
#     plt.figure(figsize=(12, 12))
#     data = df[attr]
#     if attr in ['Beta', 'Liquidity', 'Vol']:
#         data = pd.cut(df[attr], bins=bins_labels[attr]['bins'], right=False, labels=bins_labels[attr]['labels'])
#     data = pd.value_counts(data, sort=True, normalize=True)
#     labels = data.index
#
#     patches, texts, autotexts = plt.pie(data, labels=labels, colors=colors[:len(labels)],
#                                         autopct='%1.1f%%' if 1. not in data.tolist() else '%1d%%',
#                                         shadow=False, textprops={'fontsize': pie_font_size, 'color': 'w'},
#                                         pctdistance=.6 if 1. not in data.tolist() else 0.)
#     plt.title(attr)
#     plt.xticks(fontsize=xy_ticks_font_size)
#     plt.yticks(fontsize=xy_ticks_font_size)
#     plt.legend(frameon=False, fontsize=legend_font_size)
#     plt.xlabel('', fontsize=xy_label_font_size)
#     plt.ylabel('', fontsize=xy_label_font_size)
#
#     for i, autotext in enumerate(autotexts):
#         if data[i] == 0.:
#             autotext.set_visible(False)
#
#     plt.savefig(attr + '.png', bbox_inches='tight')
#
# wb = load_workbook(path)
# for attr in attributes:
#     pie = wb.create_sheet(attr)
#     pie.add_image(openpyxl.drawing.image.Image(attr + '.png'), 'A1')

# subplots
plt.figure(figsize=(12*3, 12*2))
for ind, attr in enumerate(attributes):
    data = df[attr]
    if attr in ['Beta', 'Liquidity', 'Vol']:
        data = pd.cut(df[attr], bins=bins_labels[attr]['bins'], right=False, labels=bins_labels[attr]['labels'])
    data = pd.value_counts(data, sort=True, normalize=True)
    labels = data.index
    plt.subplot(231+ind)

    patches, texts, autotexts = plt.pie(data, labels=labels, colors=colors[:len(labels)],
                                        autopct='%1.1f%%' if 1. not in data.tolist() else '%1d%%',
                                        shadow=False, textprops={'fontsize': pie_font_size, 'color': 'w'},
                                        pctdistance=.6 if 1. not in data.tolist() else 0.)

    # plt.pie(data, labels=labels, colors=colors[:len(labels)], autopct='%1.1f%%')
    plt.title(attr)
    plt.xticks(fontsize=xy_ticks_font_size)
    plt.yticks(fontsize=xy_ticks_font_size)
    plt.legend(frameon=False, fontsize=legend_font_size)
    plt.xlabel('', fontsize=xy_label_font_size)
    plt.ylabel('', fontsize=xy_label_font_size)

    for i, autotext in enumerate(autotexts):
        if data[i] == 0.:
            autotext.set_visible(False)
plt.savefig(fig_path, bbox_inches='tight')



wb = load_workbook(path)
pie = wb.create_sheet('Pie')
pie.add_image(openpyxl.drawing.image.Image(fig_path), 'A1')

wb.save(path)
